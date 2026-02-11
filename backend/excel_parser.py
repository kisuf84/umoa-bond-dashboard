"""
UMOA Yield Curve Excel Parser
Parses yield curve data from weekly Excel uploads

File structure (YC-23.janv_.26.xlsx):
- Sheet 1 "Feuil1": EMPTY - skip
- Sheets 2-9: Country data (Burkina, Bénin, Cote d'ivoire, etc.)
- Headers at Row 13: "Maturité", "Zero Coupon", "Taux Après Lissage"
- Data starts Row 14
- Columns: L (Maturité), M (Zero Coupon), N (Taux Après)
- Values are decimals (0.0984 = 9.84%)
"""

from openpyxl import load_workbook
from typing import List, Dict, Optional
from decimal import Decimal
import re


class YieldCurveExcelParser:
    """Parse yield curve Excel files with country sheets"""

    # Map sheet names to country codes
    SHEET_TO_COUNTRY = {
        'burkina': 'BF',
        'bénin': 'BJ',
        'benin': 'BJ',
        "cote d'ivoire": 'CI',
        'cote divoire': 'CI',
        'côte d\'ivoire': 'CI',
        'guinée-bissau': 'GW',
        'guinee-bissau': 'GW',
        'guinée bissau': 'GW',
        'mali': 'ML',
        'niger': 'NE',
        'sénégal': 'SN',
        'senegal': 'SN',
        'togo': 'TG',
    }

    # Maturity text to years mapping
    MATURITY_MAP = {
        '3 mois': 0.25,
        '6 mois': 0.50,
        '9 mois': 0.75,
        '1 an': 1.0,
        '2 ans': 2.0,
        '3 ans': 3.0,
        '4 ans': 4.0,
        '5 ans': 5.0,
        '6 ans': 6.0,
        '7 ans': 7.0,
        '8 ans': 8.0,
        '9 ans': 9.0,
        '10 ans': 10.0,
    }

    # Column positions (1-indexed for openpyxl)
    COL_MATURITY = 12  # Column L
    COL_ZERO_COUPON = 13  # Column M
    COL_OAT_RATE = 14  # Column N

    # Row positions
    HEADER_ROW = 13
    DATA_START_ROW = 14

    def __init__(self):
        self.errors = []
        self.warnings = []

    def parse(self, filepath: str) -> List[Dict]:
        """
        Parse Excel file and extract yield curve data for all countries.

        Returns:
            List of dicts with keys: country_code, maturity_years, zero_coupon_rate, oat_rate
        """
        self.errors = []
        self.warnings = []
        results = []

        try:
            wb = load_workbook(filepath, data_only=True)
            sheet_names = wb.sheetnames

            print(f"\n{'='*50}")
            print(f"EXCEL PARSER: Found {len(sheet_names)} sheets")
            print(f"Sheet names: {sheet_names}")
            print(f"{'='*50}")

            for sheet_name in sheet_names:
                # Skip empty/summary sheets
                sheet_name_lower = sheet_name.lower().strip()
                if sheet_name_lower in ['feuil1', 'sheet1', 'summary', 'sommaire']:
                    print(f"  Skipping sheet: {sheet_name}")
                    continue

                # Find country code from sheet name
                country_code = self._get_country_code(sheet_name)
                if not country_code:
                    self.warnings.append(f"Unknown country sheet: '{sheet_name}' - skipped")
                    print(f"  Unknown sheet: {sheet_name} - skipped")
                    continue

                sheet = wb[sheet_name]
                print(f"\n  Parsing sheet: {sheet_name} → {country_code}")

                sheet_data = self._parse_sheet(sheet, country_code, sheet_name)
                results.extend(sheet_data)

                print(f"    Extracted {len(sheet_data)} data points")

            wb.close()

            print(f"\n{'='*50}")
            print(f"TOTAL: {len(results)} yield curve points extracted")
            print(f"{'='*50}\n")

        except Exception as e:
            self.errors.append(f"Failed to parse Excel: {str(e)}")
            import traceback
            traceback.print_exc()

        return results

    def _get_country_code(self, sheet_name: str) -> Optional[str]:
        """Get country code from sheet name"""
        name_lower = sheet_name.lower().strip()

        # Direct lookup
        if name_lower in self.SHEET_TO_COUNTRY:
            return self.SHEET_TO_COUNTRY[name_lower]

        # Partial match
        for key, code in self.SHEET_TO_COUNTRY.items():
            if key in name_lower or name_lower in key:
                return code

        return None

    def _parse_sheet(self, sheet, country_code: str, sheet_name: str) -> List[Dict]:
        """Parse a single country sheet"""
        data = []

        # Read data starting from row 14
        for row_idx in range(self.DATA_START_ROW, sheet.max_row + 1):
            maturity_cell = sheet.cell(row=row_idx, column=self.COL_MATURITY).value
            zero_coupon_cell = sheet.cell(row=row_idx, column=self.COL_ZERO_COUPON).value
            oat_cell = sheet.cell(row=row_idx, column=self.COL_OAT_RATE).value

            # Skip empty rows
            if not maturity_cell:
                continue

            # Parse maturity
            maturity_years = self._parse_maturity(str(maturity_cell).strip())
            if maturity_years is None:
                continue

            # Parse rates and convert to percentage (* 100)
            zero_coupon_rate = self._parse_rate(zero_coupon_cell)
            oat_rate = self._parse_rate(oat_cell)

            # Only add if we have at least one rate
            if zero_coupon_rate is not None or oat_rate is not None:
                data.append({
                    'country_code': country_code,
                    'maturity_years': maturity_years,
                    'zero_coupon_rate': zero_coupon_rate,
                    'oat_rate': oat_rate
                })

        if not data:
            self.warnings.append(f"No data extracted from sheet '{sheet_name}' ({country_code})")

        return data

    def _parse_maturity(self, text: str) -> Optional[float]:
        """Convert maturity text to years"""
        text_lower = text.lower().strip()

        # Direct lookup
        if text_lower in self.MATURITY_MAP:
            return self.MATURITY_MAP[text_lower]

        # Try variations (remove extra spaces)
        text_normalized = ' '.join(text_lower.split())
        if text_normalized in self.MATURITY_MAP:
            return self.MATURITY_MAP[text_normalized]

        # Try partial match
        for key, value in self.MATURITY_MAP.items():
            if key in text_lower:
                return value

        # Try parsing numeric patterns like "3M", "1Y", "2A"
        match = re.match(r'(\d+)\s*(m|mois|a|an|ans|y|year|years)?', text_lower)
        if match:
            num = int(match.group(1))
            unit = match.group(2) or ''

            if unit in ['m', 'mois']:
                return round(num / 12, 2)
            elif unit in ['a', 'an', 'ans', 'y', 'year', 'years', '']:
                return float(num)

        return None

    def _parse_rate(self, value) -> Optional[float]:
        """Parse rate value and convert to percentage (multiply by 100)"""
        if value is None:
            return None

        try:
            if isinstance(value, (int, float)):
                # Values are decimals like 0.0984, convert to 9.84%
                rate = float(value) * 100
                return round(rate, 4)
            elif isinstance(value, str):
                # Remove % sign and whitespace
                cleaned = value.replace('%', '').replace(',', '.').strip()
                if cleaned:
                    rate = float(cleaned)
                    # If already looks like a percentage (> 1), don't multiply
                    if rate < 1:
                        rate = rate * 100
                    return round(rate, 4)
        except Exception:
            pass

        return None

    def get_summary(self) -> Dict:
        """Get parsing summary"""
        return {
            'errors': self.errors,
            'warnings': self.warnings
        }
